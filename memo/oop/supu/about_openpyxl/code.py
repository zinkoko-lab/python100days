from openpyxl import load_workbook

wb = load_workbook(filename="output.xlsx")
sheet_name = wb.sheetnames[0]
ws1 = wb[sheet_name]
x = ws1["A4"].value
ws2 = wb.create_sheet("new sheet")
ws2["A1"] = x
wb.save("output2.xlsx")
