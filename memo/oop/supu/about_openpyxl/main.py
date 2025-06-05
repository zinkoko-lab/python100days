from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()  # excel workbook object
ws = wb.active  # work sheet object
ws["A4"] = 10  # あるセールにデータを書き込み
cel = ws["A4"]  # セルのオブジェクト
cel.font = Font(size=12, bold=True, color="FF0000")
wb.save("output.xlsx")
