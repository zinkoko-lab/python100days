# グラフを描画
from openpyxl import load_workbook
from openpyxl.chart import Reference, LineChart

wb = load_workbook("output5.xlsx")
sheet_name = wb.sheetnames[0]
ws1 = wb[sheet_name]
x = Reference(ws1, min_col=2, min_row=2, max_col=2, max_row=13)
data = Reference(ws1, min_col=3, min_row=2, max_col=3, max_row=13)
chart = LineChart()
chart.add_data(data)
chart.set_categories(x)
ws1.add_chart(chart, "E2")
wb.save("output5.xlsx")

"""
type of charts and graphs
BarChart
LineChart
PieChart

"""
