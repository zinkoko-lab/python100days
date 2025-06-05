from openpyxl import load_workbook
from openpyxl.styles import Side, Border

wb = load_workbook("output2.xlsx")
ws1 = wb.worksheets[0]
ws1.merge_cells("A5:B7")
ws1["A5"] = "さぷー"

s = Side(style="thin")
b = Border(left=s, right=s, top=s, bottom=s)

cell = ws1["B2"]
cell.border = b
wb.save("output3.xlsx")
